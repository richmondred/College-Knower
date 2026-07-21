from __future__ import annotations

import json
import re
import subprocess
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DESKTOP = Path("/Users/tjkatsidzira/Desktop")
DOWNLOADS = Path("/Users/tjkatsidzira/Downloads")

SOURCE_FILES = {
    "cfb_champions": DESKTOP / "CFB Champions List.rtf",
    "heisman": DESKTOP / "Heisman Winners List.rtfd" / "TXT.rtf",
    "nfl_teams": DESKTOP / "NFL Teams.rtf",
    "super_bowls": DESKTOP / "Super Bowl Winners List.rtf",
    "nfl_mvp": DESKTOP / "NFL MVP List.rtf",
    "english_champions": DESKTOP / "Premier League Champions.rtf",
    "english_pyramid": DOWNLOADS / "english_football_pyramid_2026_27.xlsx",
    "europe_world": DOWNLOADS / "top_three_tiers_europe_world_2026.xlsx",
}

THEMES = {
    "cfb": {"accent": "#F43F5E", "accentHover": "#FB7185", "tint": "rgba(244, 63, 94, 0.18)"},
    "nfl": {"accent": "#22C55E", "accentHover": "#4ADE80", "tint": "rgba(34, 197, 94, 0.16)"},
    "english": {"accent": "#38BDF8", "accentHover": "#7DD3FC", "tint": "rgba(56, 189, 248, 0.15)"},
    "europe": {"accent": "#A78BFA", "accentHover": "#C4B5FD", "tint": "rgba(167, 139, 250, 0.15)"},
    "world": {"accent": "#F59E0B", "accentHover": "#FBBF24", "tint": "rgba(245, 158, 11, 0.16)"},
}


def text_from_rtf(path: Path) -> list[str]:
    output = subprocess.check_output(["textutil", "-convert", "txt", "-stdout", str(path)], text=True)
    return [line.strip() for line in output.splitlines() if line.strip()]


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKD", value) if not unicodedata.combining(char)
    )


def normalise(value: str) -> str:
    value = unicodedata.normalize("NFKC", value)
    value = re.sub(r"[\u2018\u2019\u02bb\u02bc`´]", "'", value)
    value = value.lower().replace("&", " and ")
    value = re.sub(r"[.,()/\\:\-]", " ", value)
    value = strip_accents(value)
    value = re.sub(r"\bst\b", "state", value)
    return re.sub(r"\s+", " ", value).strip()


def slug(value: str) -> str:
    value = strip_accents(value).lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "entry"


def clean_name(value: str) -> str:
    value = value.strip()
    value = re.sub(r"\[[^\]]+\]", "", value)
    value = re.sub(r"\s+\(\d+\)$", "", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip(" ,")


def decade_group(year: int) -> str:
    return f"{year // 10 * 10}s"


def entry(
    *,
    entry_id: str,
    mode_ids: list[str],
    prompt: str,
    answer: str,
    aliases: list[str],
    group: str,
    detail: str = "",
    role: str = "team",
    runner_up: bool = False,
) -> dict[str, Any]:
    values = [answer, *aliases]
    seen: set[str] = set()
    clean_aliases: list[str] = []
    for value in values:
        value = clean_name(str(value))
        key = normalise(value)
        if not value or len(key) < 2 or key in seen:
            continue
        seen.add(key)
        clean_aliases.append(value)
    return {
        "id": entry_id,
        "modeIds": mode_ids,
        "prompt": prompt,
        "answer": clean_name(answer),
        "aliases": clean_aliases,
        "group": group,
        "detail": detail,
        "role": role,
        "runnerUp": runner_up,
    }


def add_unique_last_name_aliases(entries: list[dict[str, Any]]) -> None:
    names = [entry["answer"] for entry in entries]
    last_names = []
    for name in names:
        parts = re.split(r"\s+", clean_name(name))
        last_names.append(parts[-1] if parts else "")
    counts = Counter(normalise(last) for last in last_names if last)
    for item, last in zip(entries, last_names, strict=True):
        if last and counts[normalise(last)] == 1:
            item["aliases"].append(last)


def add_unique_short_club_aliases(entries: list[dict[str, Any]]) -> None:
    candidates: dict[str, list[tuple[dict[str, Any], str]]] = defaultdict(list)
    prefixes = ("FC ", "AFC ", "AC ", "CF ", "CD ", "Club ", "FK ", "SK ", "SC ")
    suffixes = (
        " FC",
        " AFC",
        " CF",
        " C.F.",
        " S.C.",
        " SC",
        " FK",
        " SK",
        " AC",
        " CD",
        " UD",
        " IF",
        " BK",
        " United FC",
        " Football Club",
    )
    replacements = {
        "Manchester United": ["Man United", "Man Utd"],
        "Manchester City": ["Man City"],
        "Tottenham Hotspur": ["Spurs", "Tottenham"],
        "Internazionale": ["Inter"],
        "Paris Saint-Germain": ["PSG", "Paris SG"],
        "Bayern Munich": ["Bayern"],
        "Borussia Dortmund": ["Dortmund"],
        "Sporting CP": ["Sporting"],
        "Athletic Club": ["Athletic Bilbao"],
        "Atlético Madrid": ["Atletico Madrid"],
    }
    for item in entries:
        name = item["answer"]
        options = list(replacements.get(name, []))
        short = name
        for prefix in prefixes:
            if short.startswith(prefix):
                options.append(short[len(prefix) :])
        for suffix in suffixes:
            if short.endswith(suffix):
                options.append(short[: -len(suffix)])
        if " Utd" in short:
            options.append(short.replace(" Utd", " United"))
        if " St " in short or short.startswith("St "):
            options.append(short.replace("St ", "Saint "))
        for option in options:
            key = normalise(option)
            if key and key != normalise(name):
                candidates[key].append((item, option))
    for key, matches in candidates.items():
        answer_keys = {normalise(item["answer"]) for item, _ in matches}
        if len(answer_keys) == 1:
            item, alias = matches[0]
            if normalise(alias) not in {normalise(value) for value in item["aliases"]}:
                item["aliases"].append(alias)


def split_cfb_champions(value: str) -> list[str]:
    value = clean_name(value)
    if value.lower().startswith("none"):
        return []
    return [clean_name(part) for part in value.split(",") if clean_name(part)]


def cfb_aliases(team: str) -> list[str]:
    aliases = [team]
    specials = {
        "Southern California": ["USC", "Southern Cal"],
        "Miami (Fla.)": ["Miami", "Miami Florida", "The U", "Miami Hurricanes"],
        "Florida St.": ["Florida State", "FSU"],
        "Florida State": ["Florida St", "FSU"],
        "LSU": ["Louisiana State"],
        "UCLA": ["U C L A"],
        "TCU": ["Texas Christian"],
        "UCF": ["Central Florida"],
        "BYU": ["Brigham Young"],
        "Brigham Young": ["BYU"],
        "Penn State": ["Penn St"],
        "Michigan State": ["Michigan St"],
        "Ohio State": ["Ohio St"],
        "Georgia Tech": ["Georgia Institute of Technology"],
        "Texas A&M": ["Texas A and M", "Texas AM"],
        "Pittsburgh": ["Pitt"],
        "Mississippi": ["Ole Miss"],
    }
    aliases.extend(specials.get(team, []))
    return aliases


def build_cfb_champions() -> dict[str, Any]:
    lines = text_from_rtf(SOURCE_FILES["cfb_champions"])[3:]
    entries: list[dict[str, Any]] = []
    index = 0
    while index + 2 < len(lines):
        if not re.fullmatch(r"\d{4}", lines[index]):
            index += 1
            continue
        year = int(lines[index])
        champions = split_cfb_champions(lines[index + 1])
        selector = lines[index + 2]
        for champion_index, champion in enumerate(champions, start=1):
            entries.append(
                entry(
                    entry_id=f"cfb-champions-{year}-{champion_index}-{slug(champion)}",
                    mode_ids=["easy", "hard"],
                    prompt=str(year),
                    answer=champion,
                    aliases=cfb_aliases(champion),
                    group=decade_group(year),
                    detail=selector,
                    role="winner",
                )
            )
        index += 3
    return {
        "id": "cfb-past-champions",
        "title": "College Football Past Champions Quiz",
        "eyebrow": "college football knowers",
        "description": "Name the national champion selections from the supplied college football champions list.",
        "answerLabel": "Type a champion team",
        "placeholder": "Type a champion...",
        "theme": THEMES["cfb"],
        "modes": [
            {
                "id": "easy",
                "label": "Easy",
                "description": "A team fills every season it won.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "hard",
                "label": "Hard",
                "description": "Each season slot must be answered, even repeat champions.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def parse_heisman_line(lines: list[str], index: int) -> tuple[dict[str, str], int]:
    first = lines[index]
    if "," in first and not first.endswith(","):
        combined = first
        index += 1
    else:
        combined = f"{first} {lines[index + 1]}"
        index += 2
    match = re.match(r"(?P<player>.+?)\s*-\s*(?P<pos>[^,]+),\s*(?P<school>.+)$", combined)
    if not match:
        raise ValueError(f"Could not parse Heisman row: {combined!r}")
    return match.groupdict(), index


def build_heisman() -> dict[str, Any]:
    lines = text_from_rtf(SOURCE_FILES["heisman"])
    index = next(i for i, line in enumerate(lines) if re.fullmatch(r"\d{4}", line))
    entries: list[dict[str, Any]] = []
    roles = [("winner", "Winner", False), ("runner-up", "Runner-up", True), ("third", "Third place", True)]
    while index < len(lines):
        if not re.fullmatch(r"\d{4}", lines[index]):
            index += 1
            continue
        year = int(lines[index])
        index += 1
        for rank_index, (role, label, runner_up) in enumerate(roles, start=1):
            parsed, index = parse_heisman_line(lines, index)
            player = clean_name(parsed["player"])
            entries.append(
                entry(
                    entry_id=f"heisman-{year}-{rank_index}-{slug(player)}",
                    mode_ids=["easy", "hard"],
                    prompt=f"{year} {label}",
                    answer=player,
                    aliases=[player],
                    group=decade_group(year),
                    detail=f"{parsed['pos']} · {clean_name(parsed['school'])}",
                    role=role,
                    runner_up=runner_up,
                )
            )
    add_unique_last_name_aliases(entries)
    return {
        "id": "cfb-heisman-winners",
        "title": "Heisman Winners Quiz",
        "eyebrow": "college football knowers",
        "description": "Name Heisman Trophy winners, with optional runner-up and third-place rows.",
        "answerLabel": "Type a player",
        "placeholder": "Type a player...",
        "theme": THEMES["cfb"],
        "modes": [
            {
                "id": "easy",
                "label": "Easy",
                "description": "A player fills every matching Heisman row.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "hard",
                "label": "Hard",
                "description": "Each row must be answered, even repeat winners.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "runnerUpToggle": {
            "label": "Play With Runners-up",
            "description": "Adds second-place and third-place Heisman finishers.",
        },
        "entries": entries,
    }


def build_nfl_teams() -> dict[str, Any]:
    lines = [line for line in text_from_rtf(SOURCE_FILES["nfl_teams"]) if not line.startswith("Times-")]
    teams = [line for line in lines if re.match(r"^[A-Z0-9]", line)][:32]
    entries: list[dict[str, Any]] = []
    for team in teams:
        nickname = team.split()[-1]
        city = team[: -len(nickname)].strip()
        aliases = [team, nickname]
        if normalise(city) not in {"new york", "los angeles"}:
            aliases.append(city)
        entries.append(
            entry(
                entry_id=f"nfl-team-{slug(team)}",
                mode_ids=["casual", "fan", "ball-knower"],
                prompt="NFL team",
                answer=team,
                aliases=aliases,
                group=team[0].upper(),
                detail=nickname,
                role="team",
            )
        )
    return {
        "id": "nfl-all-teams",
        "title": "NFL All Teams Quiz",
        "eyebrow": "NFL knowers",
        "description": "Name all 32 NFL teams from the supplied NFL teams list.",
        "answerLabel": "Type an NFL team",
        "placeholder": "Type a team...",
        "theme": THEMES["nfl"],
        "modes": [
            {
                "id": "casual",
                "label": "Casual",
                "description": "Broad aliases and instant recognition.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "fan",
                "label": "Fan",
                "description": "Standard team-name challenge.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "ball-knower",
                "label": "Ball Knower",
                "description": "Full focus mode. Every team still counts once.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def build_nfl_champions() -> dict[str, Any]:
    lines = text_from_rtf(SOURCE_FILES["super_bowls"])[10:]
    entries: list[dict[str, Any]] = []
    index = 0
    date_pattern = re.compile(r"^[A-Z][a-z]{2} \d{1,2}, \d{4}$")
    while index < len(lines):
        if not date_pattern.match(lines[index]):
            index += 1
            continue
        row = lines[index : index + 10]
        if len(row) < 10:
            break
        date, super_bowl, winner, winner_points, loser, loser_points, mvp, *_ = row
        year = int(re.search(r"\d{4}", date).group(0))
        entries.append(
            entry(
                entry_id=f"super-bowl-{slug(super_bowl)}-{slug(winner)}",
                mode_ids=["easy", "hard"],
                prompt=super_bowl,
                answer=winner,
                aliases=[winner, winner.split()[-1]],
                group=decade_group(year),
                detail=f"{date} · beat {loser} {winner_points}-{loser_points}",
                role="winner",
            )
        )
        index += 10
    return {
        "id": "nfl-past-champions",
        "title": "Super Bowl Champions Quiz",
        "eyebrow": "NFL knowers",
        "description": "Name every Super Bowl champion from the supplied winners list.",
        "answerLabel": "Type a Super Bowl champion",
        "placeholder": "Type a champion...",
        "theme": THEMES["nfl"],
        "modes": [
            {
                "id": "easy",
                "label": "Easy",
                "description": "A team fills every Super Bowl it won.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "hard",
                "label": "Hard",
                "description": "Each Super Bowl slot must be answered, even repeat champions.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def build_nfl_mvp() -> dict[str, Any]:
    lines = text_from_rtf(SOURCE_FILES["nfl_mvp"])[6:]
    entries: list[dict[str, Any]] = []
    for index in range(0, len(lines), 6):
        row = lines[index : index + 6]
        if len(row) < 6:
            continue
        year, league, pos, player, team, _voting = row
        if not re.fullmatch(r"\d{4}", year):
            continue
        entries.append(
            entry(
                entry_id=f"nfl-mvp-{year}-{slug(player)}",
                mode_ids=["easy", "hard"],
                prompt=year,
                answer=player,
                aliases=[player],
                group=decade_group(int(year)),
                detail=f"{pos} · {team}",
                role="winner",
            )
        )
    add_unique_last_name_aliases(entries)
    return {
        "id": "nfl-mvp",
        "title": "NFL MVP Quiz",
        "eyebrow": "NFL knowers",
        "description": "Name AP NFL MVP winners from the supplied MVP list.",
        "answerLabel": "Type an MVP",
        "placeholder": "Type a player...",
        "theme": THEMES["nfl"],
        "modes": [
            {
                "id": "easy",
                "label": "Easy",
                "description": "A player fills every MVP season they won.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "hard",
                "label": "Hard",
                "description": "Each MVP season must be answered, even repeat winners.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def build_english_champions() -> dict[str, Any]:
    lines = text_from_rtf(SOURCE_FILES["english_champions"])
    entries: list[dict[str, Any]] = []
    current_group = "English champions"
    index = 0
    season_pattern = re.compile(r"^\d{4}[\u2013-]\d{2,4}$")
    while index + 4 < len(lines):
        line = lines[index]
        if "Football League" in line or "Premier League" in line:
            current_group = clean_name(line)
            index += 1
            continue
        if re.fullmatch(r"\d+", line) and season_pattern.match(lines[index + 1]):
            season = lines[index + 1]
            champion = clean_name(lines[index + 2])
            runner_up = clean_name(lines[index + 3])
            entries.append(
                entry(
                    entry_id=f"english-champion-{slug(season)}-{slug(champion)}",
                    mode_ids=["easy", "hard"],
                    prompt=f"{season} Champion",
                    answer=champion,
                    aliases=[champion],
                    group=current_group,
                    detail="Champion",
                    role="winner",
                )
            )
            if runner_up:
                entries.append(
                    entry(
                        entry_id=f"english-runner-up-{slug(season)}-{slug(runner_up)}",
                        mode_ids=["easy", "hard"],
                        prompt=f"{season} Runner-up",
                        answer=runner_up,
                        aliases=[runner_up],
                        group=current_group,
                        detail="Runner-up",
                        role="runner-up",
                        runner_up=True,
                    )
                )
            index += 5
            continue
        index += 1
    add_unique_short_club_aliases(entries)
    return {
        "id": "english-past-champions",
        "title": "English Past Champions Quiz",
        "eyebrow": "English football knowers",
        "description": "Name English top-flight champions, with optional runners-up.",
        "answerLabel": "Type a club",
        "placeholder": "Type a club...",
        "theme": THEMES["english"],
        "modes": [
            {
                "id": "easy",
                "label": "Easy",
                "description": "A club fills every season it appears.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "hard",
                "label": "Hard",
                "description": "Each season slot must be answered, even repeat clubs.",
                "solve": "single",
                "autoSubmit": True,
            },
        ],
        "runnerUpToggle": {
            "label": "Play With Runners-up",
            "description": "Adds second-place clubs for each season.",
        },
        "entries": entries,
    }


def read_sheet_rows(path: Path, sheet_name: str, header_row: int | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    if header_row is None:
        for row_number, raw in enumerate(sheet.iter_rows(min_row=1, max_row=12, values_only=True), start=1):
            values = {str(value).strip() for value in raw if value is not None}
            if "Club" in values and ("Overall tier" in values or "Rank" in values):
                header_row = row_number
                break
        else:
            raise ValueError(f"Could not find header row in {sheet_name!r}")
    headers = [str(value).strip() if value is not None else "" for value in next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    rows: list[dict[str, Any]] = []
    for raw in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        item = {headers[index]: raw[index] for index in range(min(len(headers), len(raw))) if headers[index]}
        if any(value is not None and str(value).strip() for value in item.values()):
            rows.append(item)
    workbook.close()
    return rows


def build_english_teams() -> dict[str, Any]:
    top4_rows = read_sheet_rows(SOURCE_FILES["english_pyramid"], "Top 4")
    tier7_rows = read_sheet_rows(SOURCE_FILES["english_pyramid"], "Through tier 7")
    tier9_rows = read_sheet_rows(SOURCE_FILES["english_pyramid"], "Through tier 9")
    mode_by_key: dict[str, set[str]] = defaultdict(set)
    row_by_key: dict[str, dict[str, Any]] = {}
    for mode, rows in [("top4", top4_rows), ("tier7", tier7_rows), ("tier9", tier9_rows)]:
        for row in rows:
            club = clean_name(str(row.get("Club", "")).strip())
            if not club:
                continue
            key = normalise(f"{club}::{row.get('League body')}::{row.get('Division')}")
            mode_by_key[key].add(mode)
            row_by_key[key] = row
    entries: list[dict[str, Any]] = []
    for key, row in row_by_key.items():
        club = clean_name(str(row["Club"]))
        tier = int(row.get("Overall tier") or 0)
        league = clean_name(str(row.get("League body") or ""))
        division = clean_name(str(row.get("Division") or ""))
        entries.append(
            entry(
                entry_id=f"english-club-{slug(key)}",
                mode_ids=sorted(mode_by_key[key], key=["top4", "tier7", "tier9"].index),
                prompt=f"Tier {tier}",
                answer=club,
                aliases=[club],
                group=f"Tier {tier}",
                detail=f"{league} · {division}" if division and division != league else league,
                role="club",
            )
        )
    add_unique_short_club_aliases(entries)
    entries.sort(key=lambda item: (int(item["group"].split()[-1]), item["detail"], item["answer"]))
    return {
        "id": "english-all-teams",
        "title": "English Football Teams Quiz",
        "eyebrow": "English football knowers",
        "description": "Name clubs from the supplied 2026-27 English football pyramid workbook.",
        "answerLabel": "Type a club",
        "placeholder": "Type a club...",
        "theme": THEMES["english"],
        "modes": [
            {
                "id": "top4",
                "label": "Easy",
                "description": "Premier League, Championship, League One and League Two.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "tier7",
                "label": "Medium",
                "description": "All clubs through tier 7.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "tier9",
                "label": "Hard",
                "description": "All clubs through tier 9.",
                "solve": "cascade",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def build_europe_teams() -> dict[str, Any]:
    rows = read_sheet_rows(SOURCE_FILES["europe_world"], "Europe clubs")
    entries: list[dict[str, Any]] = []
    for row in rows:
        club = clean_name(str(row.get("Club") or ""))
        if not club:
            continue
        rank = int(row.get("Rank") or 999)
        tier = int(row.get("Tier") or 0)
        association = clean_name(str(row.get("Association") or ""))
        league = clean_name(str(row.get("League body") or ""))
        division = clean_name(str(row.get("Division / group") or ""))
        modes = ["all"]
        if rank <= 10:
            modes.append("top10-all")
        if rank <= 5:
            modes.append("top5-all")
            if tier == 1:
                modes.append("top5-first")
        entries.append(
            entry(
                entry_id=f"europe-club-{rank}-{tier}-{slug(association)}-{slug(club)}",
                mode_ids=sorted(modes, key=["top5-first", "top5-all", "top10-all", "all"].index),
                prompt=f"Rank {rank} · Tier {tier}",
                answer=club,
                aliases=[club],
                group=association,
                detail=f"{league} · {division}" if division and division != league else league,
                role="club",
            )
        )
    add_unique_short_club_aliases(entries)
    entries.sort(key=lambda item: (item["modeIds"][-1], item["group"], item["detail"], item["answer"]))
    return {
        "id": "europe-all-teams",
        "title": "European Football Teams Quiz",
        "eyebrow": "European football knowers",
        "description": "Name clubs from the Europe Clubs sheet in the supplied workbook.",
        "answerLabel": "Type a club",
        "placeholder": "Type a club...",
        "theme": THEMES["europe"],
        "modes": [
            {
                "id": "top5-first",
                "label": "Easy",
                "description": "Top 5 ranked leagues, first divisions only.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "top5-all",
                "label": "Slightly Easy",
                "description": "Top 5 ranked leagues, every supplied division.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "top10-all",
                "label": "Medium",
                "description": "Top 10 ranked leagues, every supplied division.",
                "solve": "cascade",
                "autoSubmit": True,
            },
            {
                "id": "all",
                "label": "Hard",
                "description": "Every supplied European league and division.",
                "solve": "cascade",
                "autoSubmit": True,
            },
        ],
        "entries": entries,
    }


def build_world_clubs() -> dict[str, Any]:
    rows = read_sheet_rows(SOURCE_FILES["europe_world"], "World clubs")
    entries: list[dict[str, Any]] = []
    for row in rows:
        club = clean_name(str(row.get("Club") or ""))
        if not club:
            continue
        rank = int(row.get("Rank") or 999)
        tier = int(row.get("Tier") or 0)
        association = clean_name(str(row.get("Association") or ""))
        confederation = clean_name(str(row.get("Confederation") or "World"))
        league = clean_name(str(row.get("League body") or ""))
        division = clean_name(str(row.get("Division / group") or ""))
        entries.append(
            entry(
                entry_id=f"world-club-{rank}-{tier}-{slug(association)}-{slug(club)}",
                mode_ids=["ultimate"],
                prompt=f"{association} · Tier {tier}",
                answer=club,
                aliases=[club],
                group=confederation,
                detail=f"{league} · {division}" if division and division != league else league,
                role="club",
            )
        )
    add_unique_short_club_aliases(entries)
    entries.sort(key=lambda item: (item["group"], item["detail"], item["answer"]))
    return {
        "id": "world-all-clubs",
        "title": "World Football Clubs Quiz",
        "eyebrow": "World football knowers",
        "description": "The ultimate challenge: name every club from the World Clubs sheet.",
        "answerLabel": "Type a club",
        "placeholder": "Type a club...",
        "theme": THEMES["world"],
        "modes": [
            {
                "id": "ultimate",
                "label": "Ultimate",
                "description": "Every supplied club across the world.",
                "solve": "cascade",
                "autoSubmit": True,
            }
        ],
        "entries": entries,
    }


def main() -> None:
    missing = [str(path) for path in SOURCE_FILES.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("\n".join(missing))

    quizzes = {
        "cfb-past-champions": build_cfb_champions(),
        "cfb-heisman-winners": build_heisman(),
        "nfl-all-teams": build_nfl_teams(),
        "nfl-past-champions": build_nfl_champions(),
        "nfl-mvp": build_nfl_mvp(),
        "english-all-teams": build_english_teams(),
        "english-past-champions": build_english_champions(),
        "europe-all-teams": build_europe_teams(),
        "world-all-clubs": build_world_clubs(),
    }

    output = ROOT / "data" / "knowledge" / "quiz-data.ts"
    output.write_text(
        "import type { KnowledgeQuiz } from \"./types\";\n\n"
        "// Generated by scripts/generate-knowledge-data.py from the user-supplied source files.\n"
        "export const knowledgeQuizzes: Record<string, KnowledgeQuiz> = "
        + json.dumps(quizzes, ensure_ascii=False, indent=2)
        + ";\n\n"
        "export function getKnowledgeQuiz(id: string): KnowledgeQuiz {\n"
        "  const quiz = knowledgeQuizzes[id];\n"
        "  if (!quiz) throw new Error(`Unknown knowledge quiz: ${id}`);\n"
        "  return quiz;\n"
        "}\n",
        encoding="utf-8",
    )

    for quiz_id, quiz in quizzes.items():
        print(f"{quiz_id}: {len(quiz['entries'])} entries")


if __name__ == "__main__":
    main()
